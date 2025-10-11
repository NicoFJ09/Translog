import openpyxl
import os
import re
from tqdm import tqdm

def should_remove_entry(word, translation):
    """
    Check if entry should be removed based on cleanup rules.
    Returns True if entry should be REMOVED.
    """
    # Convert to lowercase for checking
    word_lower = word.lower()
    trans_lower = translation.lower()
    
    # Rule 1: Remove if "past" or "participle" anywhere in word or translation
    if 'past' in word_lower or 'past' in trans_lower:
        return True
    if 'participle' in word_lower or 'participle' in trans_lower:
        return True
    
    # Rule 2: Remove if parentheses in translation
    if '(' in translation or ')' in translation:
        return True
    
    return False

def clean_translation(translation):
    """
    Clean up the translation string.
    - Removes extra whitespace
    - Normalizes semicolons
    """
    # Split by semicolon, clean each part, rejoin
    parts = [part.strip() for part in translation.split(';')]
    # Remove empty parts
    parts = [p for p in parts if p]
    # Rejoin with semicolon and space
    return '; '.join(parts)

def cleanup_excel_file(input_file, output_file):
    """
    Clean up a single Excel file according to rules.
    Returns statistics about the cleanup.
    """
    # Load workbook
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return {'total': 0, 'removed': 0, 'kept': 0}
    if ws is None:
        print("La hoja activa no se pudo cargar correctamente.")
        return {'total': 0, 'removed': 0, 'kept': 0}
    total_rows = getattr(ws, 'max_row', 0)
    removed_count = 0
    kept_count = 0
    # Create new workbook for output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active if out_wb else None
    if out_ws is None:
        print("No se pudo crear la hoja de salida.")
        wb.close()
        return {'total': total_rows - 1, 'removed': 0, 'kept': 0}
    out_ws.title = getattr(ws, 'title', 'Sheet')
    # Copy header
    iter_rows = getattr(ws, 'iter_rows', None)
    if iter_rows is None:
        print("No se puede iterar sobre las filas: 'iter_rows' no disponible.")
        wb.close()
        out_wb.close()
        return {'total': total_rows - 1, 'removed': 0, 'kept': 0}
    header_row = next(iter_rows(min_row=1, max_row=1, values_only=True))
    if hasattr(out_ws, 'append'):
        out_ws.append(list(header_row))
    else:
        print("No se pudo escribir el encabezado en la hoja de salida.")
    # Process each row
    for row in iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        word = str(row[0]).strip() if row[0] else ""
        translation = str(row[1]).strip() if row[1] else ""
        if not word or not translation:
            continue
        # Check if should be removed
        if should_remove_entry(word, translation):
            removed_count += 1
            continue
        # Clean the translation
        clean_trans = clean_translation(translation)
        # Add to output
        if hasattr(out_ws, 'append'):
            out_ws.append([word, clean_trans])
        kept_count += 1
    # Save output file
    out_wb.save(output_file)
    wb.close()
    out_wb.close()
    return {
        'total': total_rows - 1,  # Minus header
        'removed': removed_count,
        'kept': kept_count
    }

def cleanup_all_files(input_dir):
    """
    Clean up all Excel files in the specified directory.
    """
    # Create cleaned output directory
    output_dir = os.path.join(input_dir, 'cleaned')
    os.makedirs(output_dir, exist_ok=True)
    
    # Find all Excel files
    excel_files = [f for f in os.listdir(input_dir) 
                   if f.endswith('.xlsx') and not f.startswith('~')]
    
    if not excel_files:
        print(f"⚠ No Excel files found in {input_dir}")
        return
    
    print(f"Found {len(excel_files)} Excel files to clean")
    print()
    
    # Process each file
    all_stats = {}
    
    for filename in tqdm(excel_files, desc="Cleaning files", unit="file"):
        input_path = os.path.join(input_dir, filename)
        output_path = os.path.join(output_dir, filename)
        
        try:
            stats = cleanup_excel_file(input_path, output_path)
            all_stats[filename] = stats
            
            tqdm.write(f"✓ {filename}: {stats['kept']:,} kept, {stats['removed']:,} removed")
        
        except Exception as e:
            tqdm.write(f"✗ Error processing {filename}: {e}")
    
    # Print summary
    print("\n" + "="*70)
    print("CLEANUP SUMMARY")
    print("="*70)
    print(f"{'File':<30} {'Original':<12} {'Kept':<12} {'Removed':<12}")
    print("-"*70)
    
    total_original = 0
    total_kept = 0
    total_removed = 0
    
    for filename, stats in sorted(all_stats.items()):
        print(f"{filename:<30} {stats['total']:>10,}  {stats['kept']:>10,}  {stats['removed']:>10,}")
        total_original += stats['total']
        total_kept += stats['kept']
        total_removed += stats['removed']
    
    print("-"*70)
    print(f"{'TOTAL':<30} {total_original:>10,}  {total_kept:>10,}  {total_removed:>10,}")
    print("="*70)
    
    removal_percentage = (total_removed / total_original * 100) if total_original > 0 else 0
    print(f"\nRemoval rate: {removal_percentage:.1f}%")
    print(f"\nCleaned files saved to: {output_dir}")

def export_to_prolog(input_dir, output_file='dictionary.pl'):
    """
    Export all cleaned Excel files to Prolog facts.
    Format: word(english_word, spanish_translation, category).
    """
    cleaned_dir = os.path.join(input_dir, 'cleaned')
    
    if not os.path.exists(cleaned_dir):
        print("⚠ No cleaned directory found. Run cleanup first.")
        return
    
    prolog_output = os.path.join(input_dir, output_file)
    
    excel_files = [f for f in os.listdir(cleaned_dir) 
                   if f.endswith('.xlsx') and not f.startswith('~')]
    
    print(f"\nExporting to Prolog: {prolog_output}")
    
    with open(prolog_output, 'w', encoding='utf-8') as f:
        f.write("% English-Spanish Dictionary\n")
        f.write("% Generated from cleaned Excel files\n")
        f.write("% Format: word(English, Spanish, Category).\n\n")
        
        total_facts = 0
        
        for filename in tqdm(sorted(excel_files), desc="Exporting", unit="file"):
            # Extract category from filename (e.g., dict_nouns.xlsx -> nouns)
            category = filename.replace('dict_', '').replace('.xlsx', '')
            
            f.write(f"\n% Category: {category}\n")
            
            file_path = os.path.join(cleaned_dir, filename)
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
            except Exception as e:
                print(f"Error al abrir {file_path}: {e}")
                continue
            if ws is None:
                print(f"No se pudo acceder a la hoja activa en {file_path}")
                wb.close()
                continue
            iter_rows = getattr(ws, 'iter_rows', None)
            if iter_rows is None:
                print(f"No se puede iterar sobre las filas en {file_path}")
                wb.close()
                continue
            for row in iter_rows(min_row=2, values_only=True):
                if len(row) < 2 or not row[0] or not row[1]:
                    continue
                word = str(row[0]).strip()
                translation = str(row[1]).strip()
                # Escape single quotes for Prolog
                word = word.replace("'", "\\'")
                translation = translation.replace("'", "\\'")
                # Write Prolog fact
                f.write(f"word('{word}', '{translation}', {category}).\n")
                total_facts += 1
            wb.close()
    
    print(f"✓ Exported {total_facts:,} Prolog facts to {prolog_output}")

# Main execution
if __name__ == "__main__":
    import os
    
    # Use absolute path to avoid execution issues from different locations
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(base_dir, '../data/categories')
    input_dir = os.path.abspath(input_dir)
    
    print("="*70)
    print("Dictionary Cleanup Script")
    print("="*70)
    print(f"Input directory: {input_dir}")
    print()
    
    if not os.path.exists(input_dir):
        print(f"✗ Error: Directory '{input_dir}' not found.")
        print("\nPlease ensure the directory structure is correct:")
        print("  script_location/")
        print("  └── data/")
        print("      └── categories/")
        print("          ├── dict_nouns.xlsx")
        print("          ├── dict_verbs.xlsx")
        print("          └── ...")
    else:
        # Step 1: Clean up all files
        cleanup_all_files(input_dir)
        
        # Step 2: Export to Prolog
        print("\n" + "="*70)
        export_choice = input("\nExport to Prolog format? (y/n): ").lower()
        if export_choice == 'y':
            export_to_prolog(input_dir)
        
        print("\n✓ All done!")
        print(f"\nCleaned files are in: {os.path.join(input_dir, 'cleaned')}")
