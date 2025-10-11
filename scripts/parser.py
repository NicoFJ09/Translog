import openpyxl
import re
from collections import defaultdict
from tqdm import tqdm

# Define POS categories and their identifiers
POS_CATEGORIES = {
    'nouns': ['noun', 'n'],
    'verbs': ['verb', 'v'],
    'adjectives': ['adjective', 'adj'],
    'adverbs': ['adverb', 'adv'],
    'pronouns': ['pronoun', 'pron'],
    'prepositions': ['preposition', 'prep'],
    'conjunctions': ['conjunction', 'conj'],
    'numerals': ['numeral', 'num']
}

def find_pos_tags(text):
    """
    Find all POS tags anywhere in the text.
    Returns list of found POS tags (normalized to lowercase).
    """
    found_tags = []
    text_lower = text.lower()
    
    # Check each category's identifiers
    for category, identifiers in POS_CATEGORIES.items():
        for identifier in identifiers:
            # Look for the identifier as a whole word (with word boundaries)
            # This handles: (adj), adj), (adj, "adj", etc.
            pattern = r'\b' + re.escape(identifier) + r'\b'
            if re.search(pattern, text_lower):
                found_tags.append(category)
                break  # Only add category once
    
    return found_tags

def extract_clean_translation(translation, pos_tags):
    """
    Extract clean translation text, removing POS annotations.
    """
    # Remove content in parentheses
    clean = re.sub(r'\([^)]*\)', '', translation)
    # Remove extra whitespace
    clean = ' '.join(clean.split())
    return clean.strip()

def categorize_entry(word, translation):
    """
    Categorize a word-translation pair.
    Returns dict with category as key and clean translation as value.
    Also returns whether entry was categorized.
    """
    found_pos = find_pos_tags(translation)
    
    categorized = defaultdict(list)
    
    # If no POS tags found or multiple POS tags found, goes to miscellaneous
    if len(found_pos) == 0 or len(found_pos) > 1:
        categorized['miscellaneous'].append(translation)
        return categorized, len(found_pos) == 0
    
    # Single POS tag found
    category = found_pos[0]
    clean_trans = extract_clean_translation(translation, found_pos)
    
    # Use cleaned translation if available, otherwise original
    categorized[category].append(clean_trans if clean_trans else translation)
    
    return categorized, False

def split_dictionary(input_file, output_prefix='output'):
    """
    Split dictionary Excel file into separate Excel files by POS.
    Reads from columns C (word) and D (translation).
    """
    print(f"Loading workbook: {input_file}")
    
    # Load the Excel file
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return
    if ws is None:
        print("La hoja activa no se pudo cargar correctamente.")
        return
    # Get total rows for progress bar
    total_rows = getattr(ws, 'max_row', 0)
    print(f"Total rows to process: {total_rows:,}")
    
    # Storage for each category
    category_data = defaultdict(list)
    unprocessed_entries = []  # Track entries with no POS tags
    all_processed_rows = set()  # Track which rows were processed
    
    # Process rows with progress bar
    print("\nProcessing entries...")
    iter_rows = getattr(ws, 'iter_rows', None)
    if iter_rows is None:
        print("No se puede iterar sobre las filas: 'iter_rows' no disponible.")
        return
    for idx, row in enumerate(tqdm(iter_rows(min_row=1, max_row=total_rows, values_only=True), 
                    total=total_rows, 
                    desc="Reading", 
                    unit="rows"), start=1):
        
        # Skip if not enough columns
        if len(row) < 4:
            continue
        
        # Column C (index 2) and Column D (index 3)
        word = row[2] if row[2] else ""
        translation = row[3] if row[3] else ""
        
        # Convert to string and strip
        word = str(word).strip() if word else ""
        translation = str(translation).strip() if translation else ""
        
        # Skip empty entries or header rows
        if not word or not translation or word.lower() == 'word':
            continue
        
        # Track that we processed this row
        all_processed_rows.add(idx)
        
        # Categorize the entry
        categories, no_pos = categorize_entry(word, translation)
        
        # Track entries with no POS tags
        if no_pos:
            unprocessed_entries.append((idx, word, translation))
        
        for category, clean_translations in categories.items():
            for clean_trans in clean_translations:
                category_data[category].append([word, clean_trans])
    
    wb.close()
    
    # Write separate Excel files for each category
    print("\nWriting output files...")
    all_categories = list(POS_CATEGORIES.keys()) + ['miscellaneous']
    
    for category in tqdm(all_categories, desc="Creating files", unit="files"):
        if category in category_data and category_data[category]:
            output_file = f'{output_prefix}_{category}.xlsx'
            
            # Create new workbook
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active if out_wb else None
            if out_ws is None:
                print(f"No se pudo crear la hoja para {category}")
                continue
            out_ws.title = category.capitalize()
            # Write header
            if hasattr(out_ws, 'append'):
                out_ws.append(['word', 'translation'])
                # Write sorted data
                for row_data in sorted(category_data[category]):
                    out_ws.append(row_data)
            else:
                print(f"No se pudo escribir en la hoja para {category}")
                continue
            # Save
            out_wb.save(output_file)
            out_wb.close()
            print(f"✓ Created {output_file} with {len(category_data[category]):,} entries")
    
    # Create unprocessed entries file if there are any
    if unprocessed_entries:
        print(f"\n⚠ Found {len(unprocessed_entries):,} entries with NO POS tags")
        output_file = f'{output_prefix}_NO_POS_TAG.xlsx'
        
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active if out_wb else None
        if out_ws is None:
            print("No se pudo crear la hoja para NO POS Tag")
            return
        out_ws.title = "No POS Tag"
        # Write header
        if hasattr(out_ws, 'append'):
            out_ws.append(['row_number', 'word', 'translation'])
            # Write unprocessed entries
            for row_num, word, translation in unprocessed_entries:
                out_ws.append([row_num, word, translation])
        else:
            print("No se pudo escribir en la hoja para NO POS Tag")
            return
        out_wb.save(output_file)
        out_wb.close()
        print(f"✓ Created {output_file} - REVIEW THESE ENTRIES")
    
    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    total = sum(len(data) for data in category_data.values())
    
    for category in sorted(category_data.keys()):
        count = len(category_data[category])
        percentage = (count/total*100) if total > 0 else 0
        print(f"{category:15s}: {count:6,} entries ({percentage:5.1f}%)")
    
    print("-"*60)
    print(f"{'TOTAL':15s}: {total:6,} entries")
    
    if unprocessed_entries:
        print(f"{'NO POS TAG':15s}: {len(unprocessed_entries):6,} entries (not counted in total)")
    
    print("="*60)
    
    # Verification
    print("\n" + "="*60)
    print("VERIFICATION")
    print("="*60)
    rows_in_file = total_rows - 1  # Minus header
    rows_processed = len(all_processed_rows) - 1  # Minus header if counted
    
    if rows_processed < rows_in_file:
        missing = rows_in_file - rows_processed
        print(f"⚠ WARNING: {missing} rows may not have been processed!")
        print(f"  Total rows in file: {rows_in_file:,}")
        print(f"  Rows processed: {rows_processed:,}")
    else:
        print(f"✓ All {rows_processed:,} data rows were processed")
    print("="*60)

# Main execution
if __name__ == "__main__":
    import os
    # Usar ruta absoluta para evitar problemas de ejecución desde diferentes ubicaciones
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = os.path.join(base_dir, '../data/en_es_aidict.xlsx')
    input_filename = os.path.abspath(input_filename)
    print("="*60)
    print("Dictionary POS Splitter for Excel Files")
    print("="*60)
    print(f"Input file: {input_filename}")
    print()
    if not os.path.isfile(input_filename):
        print(f"\n✗ Error: File '{input_filename}' not found.")
        data_dir = os.path.dirname(input_filename)
        print(f"Archivos disponibles en '{data_dir}':")
        try:
            files = os.listdir(data_dir)
            if not files:
                print("  (vacío)")
            for f in files:
                print(f"  - {f}")
        except Exception as e:
            print(f"No se pudo listar archivos: {e}")
        print("\nPor favor, coloca el archivo Excel en la carpeta 'data' o actualiza la ruta en el script.")
    else:
        try:
            split_dictionary(input_filename, output_prefix='dict')
            print("\n✓ All done!")
            print("\nNext steps:")
            print("1. Review dict_NO_POS_TAG.xlsx (if created)")
            print("2. Check each category file")
            print("3. Move entries from miscellaneous to proper categories if needed")
        except Exception as e:
            print(f"\n✗ Error: {e}")
            import traceback
            traceback.print_exc()